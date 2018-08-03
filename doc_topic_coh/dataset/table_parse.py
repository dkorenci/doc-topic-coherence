import openpyxl
from openpyxl import load_workbook

import re

class TableParser():
    '''
    parses xlsx table with theme labels <-> model topic links
    and creates in memory representation
    '''
    def __init__(self, table, themeCol, topicCol, dataRows):
        '''
        :param table: xlsx table file
        :param themeCol: table column with theme labels
        :param topicCol: table column with topic labels
        :param dataRows: list of (row_start, row_end) row ranges where the data is
        :return:
        '''
        self.tableFile = table;
        self.themeCol = themeCol; self.topicCol = topicCol; self.dataRows = dataRows

    def parse(self):
        '''
        parse table into graph representation
        :return: Theme2Topics
        '''
        theme2topics = Theme2Topics()
        wb = load_workbook(self.tableFile); ws = wb['Sheet1']
        for rowStart, rowEnd in self.dataRows:
            for r in range(rowStart, rowEnd+1):
                themeInd = '%s%d' % (self.themeCol, r);  theme = ws[themeInd].value
                topicInd = '%s%d' % (self.topicCol, r);  topic = ws[topicInd].value
                try: theme = self.normTheme(theme)
                except Exception as e: print 'row %s' % r; raise e
                topics = self.getTopicLabels(topic)
                for topic in topics: theme2topics.add(theme, topic)
        return theme2topics

    @staticmethod
    def normTheme(theme):
        'normalize theme label'
        return theme.strip()

    @staticmethod
    def getTopicLabels(topic):
        'split int individual topic labels'
        return [t.strip() for t in topic.split(',')]

class Theme2Topics():
    'container and helper methods for a bipartite graph between themes and topics'
    def __init__(self):
        self.themes = {}
        self.models = {}
        self.topics = {}

    def add(self, themeLabel, topicLabel):
        'add the theme and topic (and model) if they do not exist, connect them'
        theme = self.getTheme(themeLabel)
        topic = self.getTopic(topicLabel)
        theme.addTopic(topic); topic.addTheme(theme)

    def getTheme(self, label):
        'fetch or create theme by label'
        if label in self.themes: return self.themes[label]
        else:
            th = Theme(label)
            self.themes[label] = th
            return th

    def getTopic(self, label):
        'fetch or create topic by label'
        try:
            uspolFormat = re.match('~?uspolM[0-9]+\.[0-9]+', label)
            croelectFormat = re.match('~?model[0-9]+\.[0-9]+', label)
            assert uspolFormat or croelectFormat
        except AssertionError as e:
            print 'assert error: [%s] ' % label
            raise e
        clabel = TopicLabel.cleanLabel(label)
        if clabel in self.topics:
            newt = TopicLabel(label) # parse topic label
            oldt = self.topics[clabel]
            if newt.mixed:
                oldt.mixed = True
            return oldt
        else:
            t = TopicLabel(label)
            self.topics[t.label] = t
            # add to model (create model if neccesary)
            modelLabel = t.label.split('.')[0]
            model = self.getModel(modelLabel)
            model.addTopic(t)
            return t

    def getModel(self, label):
        'fetch or create model by label'
        if label in self.models: return self.models[label]
        else:
            m = ModelLabel(label)
            self.models[label] = m
            return m

class Theme():
    def __init__(self, label):
        self.label = label
        self.topics = {}
    def __eq__(self, other): return self.label == other.label
    def __hash__(self): return hash(self.label)

    def addTopic(self, topic):
        ' :param topic: TopicLabel '
        if topic.label not in self.topics:
            self.topics[topic.label] = topic

    def getTopics(self): return set(self.topics.values())

class ModelLabel():
    def __init__(self, label):
        self.label = label
        self.topics = {}
        self.index2topic = {}

    def addTopic(self, topic):
        ' :param topic: TopicLabel '
        if topic.label in self.topics: return
        self.topics[topic.label] = topic
        self.index2topic[topic.index] = topic
        topic.model = self

    def getThemes(self):
        s = set([])
        for topic in self.topics.values():
            for theme in topic.themes.values() : s.add(theme)
        return s


class TopicLabel():
    def __init__(self, label):
        if '~' in label: self.mixed = True
        else: self.mixed = False
        #print label, label.split('.')[0], label.split('.')[1]
        label = TopicLabel.cleanLabel(label)
        self.label = label
        try:
            self.index = int(label.split('.')[1])
        except ValueError as e:
            print 'index parse error: [%s]' % label
            raise e
        self.model = None
        self.themes = {}

    @staticmethod
    def cleanLabel(label): return label.strip('~')

    def __eq__(self, other): return self.label == other.label
    def __hash__(self): return hash(self.label)

    def addTheme(self, theme):
        ' :param theme: Theme '
        if theme.label not in self.themes:
            self.themes[theme.label] = theme

    def getThemes(self): return set(self.themes.values())